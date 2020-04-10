from openpyxl import load_workbook


def schedule(a, b, c):
    # a:已经下载的数据块
    # b:数据块的大小
    # c:远程文件的大小
    per = 100.0 * a * b / c
    if per > 100:
        per = 100
    print('%.2f%%' % per)


def get_mouth_list(car_data_path, column_name):
    book = load_workbook(filename=car_data_path)
    sheet = book['Sheet']

    # 口碑列表
    mouth_list = []

    # 总行数
    max_rows = sheet.max_row

    # 表头
    header = dict()
    for row in sheet.rows:
        for cell in row:
            # print('行：{0}|列：{1}|值：{2}'.format(cell.row, cell.column, cell.value))
            header[cell.value] = cell.column
        break

    # 获取需求列数据
    for j in range(1, max_rows + 1):
        if sheet.cell(row=j, column=header[column_name]).value is not None:
            mouth_list.append(sheet.cell(row=j, column=header[column_name]).value
                              .replace(' ', '').replace(',', '，'))
            # print(sheet.cell(row=2, column=26).value)

            # 显示进度
            schedule(j, 1, max_rows)

    return mouth_list


def write(mouth_path, mouth_list):
    with open(mouth_path, 'w') as f:
        f.writelines(mouth_list)

    print('写入成功！')


if __name__ == '__main__':
    # 数据源路径
    my_car_data_path = '../../../data/car_word_of_mouth.xlsx'

    # 保存路径
    my_mouth_path = '../../../data/mouth_spider/mouth.txt'

    # 需求列名
    my_column_name = 'mouth_url'

    # 获取口碑列表
    my_mouth_list = get_mouth_list(my_car_data_path, my_column_name)
    print(len(my_mouth_list))

    # 写入口碑
    # write(my_mouth_path, my_mouth_list)
