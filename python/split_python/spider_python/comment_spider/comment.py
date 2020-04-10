import requests
import json
import re
from bs4 import BeautifulSoup
from time import sleep
from random import randint
import csv
from openpyxl import load_workbook


Headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) '
                         'AppleWebKit/537.36 (KHTML, like Gecko) '
                         'Chrome/80.0.3987.106 Safari/537.36'}


def schedule(a, b, c):
    # a:已经下载的数据块
    # b:数据块的大小
    # c:远程文件的大小
    per = 100.0 * a * b / c
    if per > 100:
        per = 100
    print('%.2f%%' % per)


def get_car_list(car_data_path, column_name, finished_path):
    book = load_workbook(filename=car_data_path)
    sheet = book['Sheet']

    # 口碑列表
    car_list = []

    # 总行数
    max_rows = sheet.max_row

    # 表头
    header = dict()
    for row in sheet.rows:
        for cell in row:
            # print('行：{0}|列：{1}|值：{2}'.format(cell.row, cell.column, cell.value))
            header[cell.value] = cell.column
        break

    # 获取需求列数据
    for j in range(1, max_rows + 1):
        if sheet.cell(row=j, column=header[column_name]).value is not None:
            car_list.append(sheet.cell(row=j, column=header[column_name]).value)
            # print(sheet.cell(row=2, column=26).value)

    print('获取car_id完成！！！')

    # 去除已经爬过的car_id
    finished_id = []
    reader = csv.reader(open(finished_path, 'r'))
    for j in reader:
        finished_id.append(j[0])

    # 去重
    car_list = list(set(car_list))
    # 新的列表
    new_car_list = []
    for j in car_list:
        if str(j) not in set(finished_id):
            new_car_list.append(j)

    return new_car_list


def get_mouth_author_list(car_id):
    # 口碑页
    page = 0

    # 当前页的全部有用信息列表
    mouth_list = []

    # 作者id
    author_list = []

    while True:
        page += 1

        car_url = 'https://k.autohome.com.cn/spec/' \
                  + str(car_id) \
                  + '/index_' \
                  + str(page) + '.html'
        print('当前访问页：{0}'.format(car_url))

        # 获取网页内容
        sleep(randint(3, 8))
        html = requests.get(car_url, headers=Headers).text
        html = BeautifulSoup(html, 'lxml')

        # 获取存有口碑id的value
        mouth_input = html.findAll('input', attrs={'id': 'koubeiIdList'})[0]
        mouth_value_list = re.findall(r'value="(.+?)"', str(mouth_input))

        # 作者id
        name_pic_div = html.findAll('div', attrs={'class': 'name-pic'})
        author_href_list = []
        for j in name_pic_div:
            author_href_list.append(re.findall(r'i\.autohome\.com\.cn/(\d+)"', str(j))[0])

        # 存在口碑
        if len(mouth_value_list) != 0:
            mouth_list.extend(mouth_value_list[0].split('.'))
            author_list.extend(author_href_list)
        else:
            break

    print('汽车：{0}口碑列表获取完成！！！'.format(car_id))

    return mouth_list, author_list


def get_comment_dict(mouth_id):
    # 评论页
    page = 0

    # 当前页的全部有用信息词典
    comment_dict = dict()

    while True:
        page += 1

        # 评论url
        comment_url = 'https://reply.autohome.com.cn/ShowReply/ReplyJsonredis.ashx' \
                      '?count=10' \
                      '&page=' + str(page) + \
                      '&id=' + str(mouth_id) +  \
                      '&datatype=jsonp' \
                      '&appid=5' \
                      '&callback=jQuery'

        # 获取网页response
        html = requests.get(comment_url, headers=Headers).text
        html = re.findall(r'jQuery\((\{.+?})\)', html)[0]

        # 转换为json对象
        html_json = json.loads(html)

        # 存在评论
        if len(html_json['commentlist']) != 0:
            for k in range(0, len(html_json['commentlist'])):
                # 当前楼层
                floor = html_json['commentlist'][k]['RFloor']
                # 评论内容
                comment = html_json['commentlist'][k]['RContent'].replace(',', '，')

                # 存在回复
                if 'Quote' in html_json['commentlist'][k]:
                    # 回复楼层
                    r_floor = html_json['commentlist'][k]['Quote']['RFloor']
                else:
                    r_floor = -1

                comment_dict[floor] = [comment, r_floor]
        else:
            break

    # 显示评论总数
    print('当前口碑评论总数为{0}'.format(len(comment_dict)))

    return comment_dict


def write(car_id, mouth_id, author_id, comment_dict):
    # 保存路径
    path = '../../../../data/comment_data/comment_' \
           + str(my_car_id) \
           + '.csv'

    with open(path, 'a') as f:
        f_writer = csv.writer(f)

        # 存在评论
        if len(comment_dict) != 0:
            f_writer.writerow([car_id, mouth_id, author_id, len(comment_dict), comment_dict])
        else:
            f_writer.writerow([car_id, mouth_id, author_id, 0, ''])


def write_finish(car_id):
    with open('./finished.csv', 'a') as f:
        f_writer = csv.writer(f)

        # 已完成car_id
        f_writer.writerow([car_id])


if __name__ == '__main__':
    # 数据源路径
    my_car_data_path = '../../../../data/car_word_of_mouth.xlsx'

    # 已完成car_id路径
    my_finished_path = './finished.csv'

    # 需求列名
    my_column_name = 'type_car_id'

    # 获取口碑列表
    my_car_list = get_car_list(my_car_data_path, my_column_name, my_finished_path)

    for i in range(1, len(my_car_list)):
        # 汽车id
        my_car_id = my_car_list[i]

        # 获得口碑列表
        try:
            my_mouth_list, my_author_list = get_mouth_author_list(my_car_id)
        except:
            print('获得口碑列表异常！！！')
            continue

        # 获取评论词典
        for ii in range(0, len(my_mouth_list)):
            try:
                my_comment_dict = get_comment_dict(my_mouth_list[ii])
            except:
                print('获取评论词典异常！！！')
                continue

            # 保存
            write(my_car_id, my_mouth_list[ii], my_author_list[ii], my_comment_dict)

        # finished文件
        write_finish(my_car_id)
